from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pdfplumber
import re
import io
import sqlite3
import hashlib
import jwt
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = FastAPI(title="BrokerNote AI", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SECRET_KEY = "brokernote-ai-secret-2025"
ALGORITHM = "HS256"


# ─── Database ────────────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect('/tmp/brokernote.db')
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            firm TEXT,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ca_clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            pan TEXT,
            mobile TEXT NOT NULL,
            filing_type TEXT NOT NULL,
            due_date TEXT NOT NULL,
            status TEXT DEFAULT 'Pending',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ca_message_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            client_name TEXT,
            filing_type TEXT,
            due_date TEXT,
            mobile TEXT,
            sent_at TEXT,
            result TEXT
        );
    ''')
    conn.commit()
    conn.close()

init_db()


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def create_token(email: str) -> str:
    payload = {
        'email': email,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(days=30)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def verify_token(token: str) -> str:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload['email']
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


# ─── Auth ────────────────────────────────────────────────────────────────────

@app.get("/")
async def root():
    return {"status": "ok", "service": "BrokerNote AI", "version": "2.0.0"}


@app.post("/auth/signup")
async def signup(request_data: dict):
    email = request_data.get('email', '').lower().strip()
    password = request_data.get('password', '')
    name = request_data.get('name', '').strip()
    firm = request_data.get('firm', '').strip()

    if not email or not password or not name:
        raise HTTPException(status_code=400, detail="Email, password and name are required")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    conn = sqlite3.connect('/tmp/brokernote.db')
    try:
        conn.execute(
            "INSERT INTO users (email, password, name, firm, created_at) VALUES (?,?,?,?,?)",
            (email, hash_password(password), name, firm, datetime.datetime.now().isoformat())
        )
        conn.commit()
        token = create_token(email)
        return {"token": token, "email": email, "name": name, "firm": firm}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Email already registered")
    finally:
        conn.close()


@app.post("/auth/login")
async def login(request_data: dict):
    email = request_data.get('email', '').lower().strip()
    password = request_data.get('password', '')

    conn = sqlite3.connect('/tmp/brokernote.db')
    user = conn.execute(
        "SELECT email, name, firm FROM users WHERE email=? AND password=?",
        (email, hash_password(password))
    ).fetchone()
    conn.close()

    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")

    token = create_token(user[0])
    return {"token": token, "email": user[0], "name": user[1], "firm": user[2]}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        cleaned = str(val).replace(',', '').replace(' ', '').strip()
        cleaned = re.sub(r'\((.+)\)', r'-\1', cleaned)
        return float(cleaned)
    except Exception:
        return 0.0


def extract_header(text: str) -> dict:
    header = {}
    patterns = {
        "client_name":      r"Client Name\s*:\s*(.+?)(?:\n|Trading|Address)",
        "client_code":      r"Client Code\s*\(UCC\)\s*:\s*(\S+)",
        "pan":              r"PAN Number\s*:\s*([A-Z]{5}[0-9]{4}[A-Z])",
        "contract_note_no": r"Contract Note No\s*:\s*(\S+)",
        "trade_date":       r"Trade Date\s*:\s*([\d/]+)",
        "settlement_no":    r"SETTLEMENT NO\s+(\S+)",
        "gst_no":           r"GST NO\s*:\s*([A-Z0-9]+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if match:
            header[key] = match.group(1).strip()
    return header


def extract_broker_info(text: str) -> dict:
    info = {}
    brokers = ['Angel One', 'Zerodha', 'Upstox', 'HDFC Securities', 'ICICI Direct',
               'Sharekhan', 'Kotak Securities', 'Motilal Oswal', '5Paisa', 'Groww']
    for broker in brokers:
        if broker.lower() in text.lower():
            info["broker"] = broker
            break
    gst_match = re.search(r'GST NO\.?:?\s*([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][A-Z0-9]Z[A-Z0-9])', text)
    if gst_match:
        info["broker_gst"] = gst_match.group(1)
    pan_match = re.search(r'PAN No\.?:?\s*([A-Z]{5}[0-9]{4}[A-Z])', text)
    if pan_match:
        info["broker_pan"] = pan_match.group(1)
    return info


# ─── Equity Trades ───────────────────────────────────────────────────────────

def extract_equity_from_table(table: list) -> list:
    """Extract equity (ISIN-based) trades from a table."""
    trades = []
    is_trade_table = False
    for row in table:
        if not row:
            continue
        row_str = ' '.join([str(c) for c in row if c])
        if 'ISIN' in row_str and ('BUY' in row_str or 'Security' in row_str):
            is_trade_table = True
            continue
        if is_trade_table and row[0] and str(row[0]).startswith('INE'):
            try:
                trades.append({
                    "isin":                    (row[0] or '').strip(),
                    "security":                (row[1] or '').strip().replace('\n', ' '),
                    "buy_qty":                 safe_float(row[2] if len(row) > 2 else 0),
                    "buy_wap":                 safe_float(row[3] if len(row) > 3 else 0),
                    "total_buy_value":         safe_float(row[6] if len(row) > 6 else 0),
                    "sell_qty":                safe_float(row[7] if len(row) > 7 else 0),
                    "sell_wap":                safe_float(row[8] if len(row) > 8 else 0),
                    "total_sell_value":        safe_float(row[11] if len(row) > 11 else 0),
                    "net_qty":                 safe_float(row[12] if len(row) > 12 else 0),
                    "net_obligation":          safe_float(row[13] if len(row) > 13 else 0),
                })
            except Exception:
                pass
    return trades


# ─── Derivative Trades ────────────────────────────────────────────────────────

def extract_derivatives_from_table(table: list) -> list:
    """Extract F&O / derivative trades from a table."""
    derivatives = []
    is_deriv_table = False

    DERIV_KEYWORDS = ['FUTSTK', 'FUTIDX', 'OPTIDX', 'OPTSTK', 'FUTCOM', 'FUTSTK']

    for row in table:
        if not row:
            continue
        row_str = ' '.join([str(c) for c in row if c])

        # Detect derivative header
        if 'Contract Description' in row_str and ('Buy' in row_str or 'Sell' in row_str):
            is_deriv_table = True
            continue

        if is_deriv_table and row[0]:
            first_cell = str(row[0]).strip()
            # Skip total row
            if 'Total' in first_cell or 'TOTAL' in first_cell:
                continue
            # Check if it's a derivative row (contract name contains known types)
            has_deriv_keyword = any(kw in first_cell.upper() for kw in DERIV_KEYWORDS)
            has_may_jun = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{2}\s+\d{4}', first_cell)
            if has_deriv_keyword or has_may_jun:
                try:
                    # Determine instrument type
                    if 'FUTSTK' in first_cell.upper() or 'FUTIDX' in first_cell.upper():
                        instrument = 'Futures'
                    elif 'OPTIDX' in first_cell.upper() or 'OPTSTK' in first_cell.upper():
                        instrument = 'Options'
                    else:
                        instrument = 'Derivatives'

                    derivatives.append({
                        "contract":      first_cell.replace('\n', ' '),
                        "instrument":    instrument,
                        "buy_sell":      str(row[1] or '').strip().replace('\n', ' '),
                        "qty":           safe_float(row[2] if len(row) > 2 else 0),
                        "wap":           safe_float(row[3] if len(row) > 3 else 0),
                        "wap_after_brokerage": safe_float(row[5] if len(row) > 5 else 0),
                        "closing_rate":  safe_float(row[6] if len(row) > 6 else 0),
                        "net_total":     safe_float(row[7] if len(row) > 7 else 0),
                    })
                except Exception:
                    pass

    return derivatives


# ─── Charges ─────────────────────────────────────────────────────────────────

def extract_charges_from_table(table: list) -> dict:
    """Extract consolidated charges from TOTAL(NET) row.
    Returns category-wise totals with GST = CGST + SGST combined."""
    for row in table:
        if not row:
            continue
        row_str = ' '.join([str(c) for c in row if c])
        if 'TOTAL' in row_str and len([c for c in row if c]) >= 8:
            try:
                nums = [safe_float(c) for c in row
                        if c and str(c).strip() not in ('TOTAL(NET)', 'TOTAL', 'NSE-CAPITAL',
                                                         'NSE-FUTURES', 'BSE-CAPITAL')]
                if len(nums) >= 8:
                    cgst = nums[3] if len(nums) > 3 else 0
                    sgst = nums[4] if len(nums) > 4 else 0
                    return {
                        "pay_obligation":   nums[0],
                        "stt":              nums[1],
                        "taxable_value":    nums[2],   # brokerage taxable value
                        "cgst":             cgst,
                        "sgst":             sgst,
                        "gst":              round(cgst + sgst, 4),   # combined GST
                        "exchange_charges": nums[5] if len(nums) > 5 else 0,
                        "sebi_fees":        nums[6] if len(nums) > 6 else 0,
                        "stamp_duty":       nums[7] if len(nums) > 7 else 0,
                        "ipf_charges":      nums[8] if len(nums) > 8 else 0,
                        "net_amount":       nums[-1],
                    }
            except Exception:
                pass
    return {}


# ─── Order Details ────────────────────────────────────────────────────────────

def extract_orders_from_table(table: list) -> list:
    orders = []
    is_order_table = False
    for row in table:
        if not row:
            continue
        row_str = ' '.join([str(c) for c in row if c])
        if 'Order No' in row_str and 'Trade No' in row_str:
            is_order_table = True
            continue
        if is_order_table and row[0] and re.match(r'\d{10,}', str(row[0]).strip()):
            try:
                orders.append({
                    "order_no":   str(row[0]).strip(),
                    "order_time": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                    "trade_no":   str(row[2]).strip() if len(row) > 2 and row[2] else "",
                    "trade_time": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                    "security":   str(row[4]).strip().replace('\n', ' ') if len(row) > 4 and row[4] else "",
                    "buy_sell":   str(row[5]).strip() if len(row) > 5 and row[5] else "",
                    "qty":        str(row[6]).strip() if len(row) > 6 and row[6] else "",
                    "gross_rate": str(row[11]).strip() if len(row) > 11 and row[11] else "",
                })
            except Exception:
                pass
    return orders


# ─── Main Extraction ──────────────────────────────────────────────────────────

def extract_contract_note(pdf_bytes: bytes) -> dict:
    result = {
        "header":        {},
        "broker_info":   {},
        "trades":        [],    # equity trades (ISIN-based)
        "derivatives":   [],    # F&O / derivative trades
        "charges":       {},
        "order_details": [],
    }
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text  = page.extract_text() or ""
            tables = page.extract_tables()

            if page_num == 0:
                result["header"]      = extract_header(text)
                result["broker_info"] = extract_broker_info(text)

                for table in tables:
                    # Equity trades
                    eq = extract_equity_from_table(table)
                    if eq:
                        result["trades"].extend(eq)

                    # Derivative trades
                    dv = extract_derivatives_from_table(table)
                    if dv:
                        result["derivatives"].extend(dv)

                    # Charges
                    ch = extract_charges_from_table(table)
                    if ch:
                        result["charges"] = ch

            elif page_num == 1:
                for table in tables:
                    od = extract_orders_from_table(table)
                    if od:
                        result["order_details"].extend(od)

    return result


@app.post("/extract")
async def extract_pdf(file: UploadFile = File(...)):
    if not (file.filename or '').lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Please upload a PDF file")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    try:
        return extract_contract_note(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")


# ─── Excel Styles ─────────────────────────────────────────────────────────────

DARK_BLUE  = "1B2A4A"
MID_BLUE   = "2563EB"
PURPLE     = "7C3AED"
LIGHT_BLUE = "EBF3FF"
ALT_ROW    = "F5F8FF"
WHITE      = "FFFFFF"
GREEN_DARK = "14532D"
GREEN_MID  = "16A34A"

def make_border():
    side = Side(style='thin', color='D0D7E5')
    return Border(left=side, right=side, top=side, bottom=side)

def hdr_cell(cell, text, color=MID_BLUE):
    cell.value = text
    cell.font  = Font(bold=True, color=WHITE, size=10)
    cell.fill  = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = make_border()

def title_merge(ws, cell_range, text, color=LIGHT_BLUE, text_color=DARK_BLUE):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(':')[0]]
    c.value = text
    c.font  = Font(bold=True, color=text_color, size=13)
    c.fill  = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')

def section_title(ws, row, col, text, color=DARK_BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=11, color=color)
    c.fill = PatternFill("solid", fgColor="E8F0FE")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 9)


# ─── Excel Generation ─────────────────────────────────────────────────────────

@app.post("/excel")
async def generate_excel(data: dict):
    wb = openpyxl.Workbook()

    header_info   = data.get('header', {})
    broker_info   = data.get('broker_info', {})
    charges       = data.get('charges', {})
    equity_trades = data.get('trades', [])
    derivatives   = data.get('derivatives', [])
    order_details = data.get('order_details', [])

    # ── Sheet 1: Full Trade Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Trade Summary"
    ws.sheet_view.showGridLines = False

    # Title
    title_merge(ws, 'A1:N1', '📊  CONTRACT NOTE — FULL TRADE SUMMARY  |  BrokerNote AI')
    ws.row_dimensions[1].height = 32

    # Client info
    info_pairs = [
        ("Client Name",       header_info.get('client_name', '—')),
        ("Client Code (UCC)", header_info.get('client_code', '—')),
        ("PAN Number",        header_info.get('pan', '—')),
        ("Trade Date",        header_info.get('trade_date', '—')),
        ("Contract Note No.", header_info.get('contract_note_no', '—')),
        ("Settlement No.",    header_info.get('settlement_no', '—')),
        ("Broker",            broker_info.get('broker', '—')),
        ("Broker GST",        broker_info.get('broker_gst', '—')),
    ]
    for i, (label, value) in enumerate(info_pairs):
        r = 3 + i
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = Font(bold=True, size=10, color=DARK_BLUE)
        vc.font = Font(size=10)
        lc.border = vc.border = make_border()
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20

    current_row = 13

    # ── Equity Section ───────────────────────────────────────────────────────
    if equity_trades:
        section_title(ws, current_row, 1, f"  📈  EQUITY SEGMENT — {len(equity_trades)} Securities")
        current_row += 1

        eq_cols = [
            ('A', 'ISIN', 14), ('B', 'Security Name', 22),
            ('C', 'Buy Qty', 10), ('D', 'Buy WAP (₹)', 14), ('E', 'Total Buy Value (₹)', 18),
            ('F', 'Sell Qty', 10), ('G', 'Sell WAP (₹)', 14), ('H', 'Total Sell Value (₹)', 18),
            ('I', 'Net Qty', 10), ('J', 'Net Obligation (₹)', 18),
        ]
        for col_letter, col_name, width in eq_cols:
            hdr_cell(ws[f'{col_letter}{current_row}'], col_name, MID_BLUE)
            ws.column_dimensions[col_letter].width = width
        current_row += 1

        eq_buy_total = eq_sell_total = 0
        for i, trade in enumerate(equity_trades):
            fill = PatternFill("solid", fgColor=ALT_ROW) if i % 2 == 0 else None
            vals = [
                trade.get('isin', ''),        trade.get('security', ''),
                trade.get('buy_qty', 0),       trade.get('buy_wap', 0),
                trade.get('total_buy_value', 0),
                trade.get('sell_qty', 0),      trade.get('sell_wap', 0),
                trade.get('total_sell_value', 0),
                trade.get('net_qty', 0),       trade.get('net_obligation', 0),
            ]
            eq_buy_total  += vals[4]
            eq_sell_total += vals[7]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.border = make_border()
                c.font   = Font(size=10)
                if fill: c.fill = fill
                if col > 2 and isinstance(val, float):
                    c.number_format = '#,##0.00'
                if col in (3, 6, 9):
                    c.number_format = '#,##0'
            current_row += 1

        # Equity subtotal
        for col in range(1, 11):
            c = ws.cell(row=current_row, column=col)
            c.fill   = PatternFill("solid", fgColor="DBEAFE")
            c.border = make_border()
            c.font   = Font(bold=True, size=10)
        ws.cell(row=current_row, column=1, value="EQUITY TOTAL").font = Font(bold=True, size=10, color=DARK_BLUE)
        ws.cell(row=current_row, column=5, value=eq_buy_total).number_format  = '#,##0.00'
        ws.cell(row=current_row, column=5).font = Font(bold=True, color=GREEN_MID)
        ws.cell(row=current_row, column=8, value=eq_sell_total).number_format = '#,##0.00'
        ws.cell(row=current_row, column=8).font = Font(bold=True, color=GREEN_MID)
        current_row += 2

    # ── Derivative Section ───────────────────────────────────────────────────
    if derivatives:
        section_title(ws, current_row, 1, f"  📊  DERIVATIVE SEGMENT — {len(derivatives)} Contracts")
        current_row += 1

        dv_cols = [
            ('A', 'Contract Description', 36), ('B', 'Instrument', 12),
            ('C', 'B/S', 8), ('D', 'Qty', 8),
            ('E', 'WAP / Unit (₹)', 16), ('F', 'WAP After Brokerage (₹)', 22),
            ('G', 'Closing Rate (₹)', 18), ('H', 'Net P&L (₹)', 16),
        ]
        for col_letter, col_name, width in dv_cols:
            hdr_cell(ws[f'{col_letter}{current_row}'], col_name, PURPLE)
            ws.column_dimensions[col_letter].width = width
        current_row += 1

        net_pnl = 0
        for i, deriv in enumerate(derivatives):
            fill = PatternFill("solid", fgColor="F5F3FF") if i % 2 == 0 else None
            bs = deriv.get('buy_sell', '')
            net = deriv.get('net_total', 0)
            net_pnl += net
            vals = [
                deriv.get('contract', ''), deriv.get('instrument', ''),
                bs, deriv.get('qty', 0),
                deriv.get('wap', 0), deriv.get('wap_after_brokerage', 0),
                deriv.get('closing_rate', 0), net,
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.border = make_border()
                c.font   = Font(size=10)
                if fill: c.fill = fill
                if col == 3:  # B/S color
                    c.font = Font(size=10, bold=True,
                                  color="16A34A" if 'BUY' in str(val).upper() else "DC2626")
                if col >= 5 and isinstance(val, float):
                    c.number_format = '#,##0.00'
                if col == 4:
                    c.number_format = '#,##0'
                if col == 8 and isinstance(val, float):
                    c.font = Font(bold=True, color="16A34A" if val >= 0 else "DC2626")
            current_row += 1

        # Derivative net P&L subtotal
        for col in range(1, 9):
            c = ws.cell(row=current_row, column=col)
            c.fill   = PatternFill("solid", fgColor="EDE9FE")
            c.border = make_border()
            c.font   = Font(bold=True, size=10)
        ws.cell(row=current_row, column=1, value="DERIVATIVE NET P&L").font = Font(bold=True, size=10, color=PURPLE)
        npnl_cell = ws.cell(row=current_row, column=8, value=round(net_pnl, 2))
        npnl_cell.number_format = '#,##0.00'
        npnl_cell.font = Font(bold=True, color="16A34A" if net_pnl >= 0 else "DC2626")
        current_row += 2

    # ── Charges Section ──────────────────────────────────────────────────────
    section_title(ws, current_row, 1, "  💰  CHARGES & EXPENSES — Category Wise Total (All Segments)")
    current_row += 1

    gst   = charges.get('gst', round(charges.get('cgst', 0) + charges.get('sgst', 0), 2))
    stt   = charges.get('stt', 0)
    exc   = charges.get('exchange_charges', 0)
    sebi  = charges.get('sebi_fees', 0)
    stamp = charges.get('stamp_duty', 0)
    ipf   = charges.get('ipf_charges', 0)
    brok_tv = charges.get('taxable_value', 0)   # brokerage taxable value
    net   = charges.get('net_amount', 0)
    pay_ob= charges.get('pay_obligation', 0)

    charge_rows = [
        ("Gross Obligation (Pay / Receive)",           pay_ob,   False),
        ("Securities Transaction Tax (STT)",           stt,      False),
        ("Brokerage (Taxable Value of Supply)",        brok_tv,  False),
        ("GST Expense  (CGST @ 9% + SGST @ 9%)",      gst,      False),
        ("Exchange Transaction Charges",               exc,      False),
        ("SEBI Turnover Fees",                         sebi,     False),
        ("Stamp Duty",                                 stamp,    False),
        ("IPF Charges",                                ipf,      False),
        ("NET AMOUNT RECEIVABLE / (PAYABLE)",          net,      True),
    ]

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 20

    for label, value, is_total in charge_rows:
        if value == 0 and not is_total:
            current_row += 1
            continue
        lc = ws.cell(row=current_row, column=1, value=label)
        vc = ws.cell(row=current_row, column=2, value=abs(value) if is_total else value)
        lc.border = vc.border = make_border()
        vc.number_format = '₹#,##0.00'
        if is_total:
            for c in [lc, vc]:
                c.fill = PatternFill("solid", fgColor=DARK_BLUE)
                c.font = Font(bold=True, color=WHITE, size=11)
            # Add receivable/payable note
            note_cell = ws.cell(row=current_row, column=3,
                                value="← Receivable" if value >= 0 else "← Payable")
            note_cell.font = Font(italic=True, color="16A34A" if value >= 0 else "DC2626", size=10)
        else:
            lc.font = Font(size=10, color=DARK_BLUE)
            vc.font = Font(size=10)
        current_row += 1

    trade_date = header_info.get('trade_date', str(datetime.date.today()))
    broker_name = broker_info.get('broker', 'Broker')

    # ── Build Dr / Cr ledger lists (used for both Excel Sheet 2 & Tally XML) ──
    # Convention: dr_entries / cr_entries = list of (ledger_name, amount, narration)

    dr_entries = []
    cr_entries = []

    # Equity buys → Dr. Stock A/c
    for trade in equity_trades:
        buy_val = trade.get('total_buy_value', 0)
        if buy_val > 0:
            sec = trade.get('security', 'Investment')
            narr = (f"Purchase: {sec} × {int(trade.get('buy_qty',0))} "
                    f"shares @ ₹{trade.get('buy_wap',0):.2f}")
            dr_entries.append((f"{sec} A/c", buy_val, narr))

    # Equity sells → Cr. Stock A/c
    for trade in equity_trades:
        sell_val = trade.get('total_sell_value', 0)
        if sell_val > 0:
            sec = trade.get('security', 'Investment')
            narr = (f"Sale: {sec} × {int(trade.get('sell_qty',0))} "
                    f"shares @ ₹{trade.get('sell_wap',0):.2f}")
            cr_entries.append((f"{sec} A/c", sell_val, narr))

    # F&O: loss → Dr. F&O P&L A/c | profit → Cr. F&O P&L A/c
    for deriv in derivatives:
        net = deriv.get('net_total', 0)
        if net == 0:
            continue
        contract_short = deriv.get('contract', 'F&O Contract')[:40]
        narr = f"{deriv.get('buy_sell','')} {contract_short}"
        if net < 0:
            dr_entries.append(('F&O Trading P&L A/c', abs(net), narr))
        else:
            cr_entries.append(('F&O Trading P&L A/c', net, narr))

    # Expenses → Dr.
    expense_entries = [
        ('STT Expense A/c',          stt,   'Securities Transaction Tax (STT)'),
        ('GST Expense A/c',          gst,   'GST on Brokerage  (CGST @ 9% + SGST @ 9%)'),
        ('Exchange Charges Exp A/c', exc,   'Exchange Transaction Charges'),
        ('SEBI Charges Exp A/c',     sebi,  'SEBI Turnover Fees'),
        ('Stamp Duty Expense A/c',   stamp, 'Stamp Duty'),
        ('IPF Charges Exp A/c',      ipf,   'IPF / Investor Protection Fund'),
    ]
    for ledger, amt, narr in expense_entries:
        if amt and amt > 0:
            dr_entries.append((ledger, amt, narr))

    # Net balancing entry: broker payable / receivable
    total_dr_val = sum(a for _, a, _ in dr_entries)
    total_cr_val = sum(a for _, a, _ in cr_entries)
    net_diff = round(total_dr_val - total_cr_val, 2)   # positive → payable, negative → receivable

    if net_diff > 0:
        cr_entries.append((f'{broker_name} Payable A/c', net_diff,
                           f'Net amount payable to {broker_name}'))
    elif net_diff < 0:
        dr_entries.append((f'{broker_name} Receivable A/c', abs(net_diff),
                           f'Net amount receivable from {broker_name}'))

    # ── Sheet 2: Tally Journal Entry ─────────────────────────────────────────
    ws2 = wb.create_sheet("Tally Journal Entry")
    ws2.sheet_view.showGridLines = False
    title_merge(ws2, 'A1:F1', '📒  TALLY JOURNAL ENTRY  |  Ready to import in TallyPrime  |  BrokerNote AI')
    ws2.row_dimensions[1].height = 32

    jcols = [
        ('A', 'Date', 14), ('B', 'Dr / Cr', 8),
        ('C', 'Ledger Name (create in Tally if missing)', 40),
        ('D', 'Narration / Particulars', 46),
        ('E', 'Debit (₹)', 16), ('F', 'Credit (₹)', 16),
    ]
    for col_letter, col_name, width in jcols:
        hdr_cell(ws2[f'{col_letter}3'], col_name)
        ws2.column_dimensions[col_letter].width = width

    journal_rows = []
    # Dr entries
    for ledger, amt, narr in dr_entries:
        journal_rows.append([trade_date, 'Dr', ledger, narr, amt, ''])
    # Cr entries
    for ledger, amt, narr in cr_entries:
        journal_rows.append([trade_date, 'Cr', ledger, narr, '', amt])

    j_start = 4
    for i, jr in enumerate(journal_rows):
        r = j_start + i
        fill = PatternFill("solid", fgColor=ALT_ROW) if i % 2 == 0 else None
        for col, val in enumerate(jr, 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.border = make_border()
            c.font   = Font(size=10)
            if fill: c.fill = fill
            if col == 2:   # Dr/Cr label
                c.font = Font(bold=True, color='1D4ED8' if val == 'Dr' else 'DC2626', size=10)
            if col in (5, 6) and isinstance(val, (int, float)) and val:
                c.number_format = '#,##0.00'

    # Totals row
    total_row = j_start + len(journal_rows)
    for col in range(1, 7):
        c = ws2.cell(row=total_row, column=col)
        c.fill   = PatternFill("solid", fgColor=DARK_BLUE)
        c.border = make_border()
        c.font   = Font(bold=True, color=WHITE, size=10)
    ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, color=WHITE)
    tot_dr = sum(a for _, a, _ in dr_entries)
    tot_cr = sum(a for _, a, _ in cr_entries)
    drc = ws2.cell(row=total_row, column=5, value=round(tot_dr, 2))
    drc.number_format = '#,##0.00'
    drc.font = Font(bold=True, color=WHITE)
    crc = ws2.cell(row=total_row, column=6, value=round(tot_cr, 2))
    crc.number_format = '#,##0.00'
    crc.font = Font(bold=True, color=WHITE)

    # Balance check note
    bal_row = total_row + 1
    bal_ok = abs(tot_dr - tot_cr) < 0.05
    note = ws2.cell(row=bal_row, column=5,
                    value='✅ Balanced' if bal_ok else f'⚠ Diff: {round(tot_dr - tot_cr, 2)}')
    note.font = Font(bold=True, color='16A34A' if bal_ok else 'DC2626', size=10)
    ws2.merge_cells(start_row=bal_row, start_column=5, end_row=bal_row, end_column=6)

    # Ledger setup guide
    guide_row = bal_row + 3
    guide_hdr = ws2.cell(row=guide_row, column=1,
                         value='📋  TALLY LEDGER SETUP GUIDE — Create these ledgers before importing XML')
    guide_hdr.font = Font(bold=True, size=11, color=DARK_BLUE)
    guide_hdr.fill = PatternFill("solid", fgColor="E8F0FE")
    ws2.merge_cells(start_row=guide_row, start_column=1, end_row=guide_row, end_column=6)

    ledger_guide = [
        ('Ledger Name',                  'Group in Tally',         'Notes'),
        ('Each Stock A/c (e.g. Infosys A/c)', 'Investments',       'One ledger per security'),
        ('F&O Trading P&L A/c',          'Indirect Income/Loss',   'For F&O net profit/loss'),
        (f'{broker_name} Payable A/c',   'Sundry Creditors',       'Broker net payable'),
        (f'{broker_name} Receivable A/c','Sundry Debtors',         'Broker net receivable'),
        ('STT Expense A/c',              'Indirect Expenses',      'Securities Transaction Tax'),
        ('GST Expense A/c',              'Duties & Taxes',         'CGST + SGST on brokerage'),
        ('Exchange Charges Exp A/c',     'Indirect Expenses',      'NSE / BSE exchange fees'),
        ('SEBI Charges Exp A/c',         'Indirect Expenses',      'SEBI turnover fees'),
        ('Stamp Duty Expense A/c',       'Indirect Expenses',      'State stamp duty'),
    ]
    for gi, (ln, grp, note_txt) in enumerate(ledger_guide):
        gr = guide_row + 1 + gi
        cells = [
            ws2.cell(row=gr, column=1, value=ln),
            ws2.cell(row=gr, column=2, value=grp),
            ws2.cell(row=gr, column=3, value=note_txt),
        ]
        for gc in cells:
            gc.border = make_border()
            gc.font   = Font(bold=(gi == 0), size=10,
                             color=DARK_BLUE if gi == 0 else '000000')
            if gi == 0:
                gc.fill = PatternFill("solid", fgColor="DBEAFE")

    # ── Sheet 3: Order Details ────────────────────────────────────────────────
    if order_details:
        ws3 = wb.create_sheet("Order Details")
        ws3.sheet_view.showGridLines = False
        title_merge(ws3, 'A1:H1', '🗂  ORDER & TRADE DETAILS  |  BrokerNote AI')
        ws3.row_dimensions[1].height = 32
        od_cols = [
            ('A', 'Order No', 22), ('B', 'Order Time', 14), ('C', 'Trade No', 22),
            ('D', 'Trade Time', 14), ('E', 'Security / Contract', 26),
            ('F', 'B/S', 8), ('G', 'Qty', 10), ('H', 'Gross Rate (₹)', 18),
        ]
        for col_letter, col_name, width in od_cols:
            hdr_cell(ws3[f'{col_letter}3'], col_name)
            ws3.column_dimensions[col_letter].width = width
        for i, order in enumerate(order_details):
            r = 4 + i
            vals = [order.get('order_no'), order.get('order_time'), order.get('trade_no'),
                    order.get('trade_time'), order.get('security'), order.get('buy_sell'),
                    order.get('qty'), order.get('gross_rate')]
            fill = PatternFill("solid", fgColor=ALT_ROW) if i % 2 == 0 else None
            for col, val in enumerate(vals, 1):
                c = ws3.cell(row=r, column=col, value=val)
                c.border = make_border()
                c.font   = Font(size=10)
                if fill: c.fill = fill

    # ── Sheet 4: P&L Summary ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("P&L Summary")
    ws4.sheet_view.showGridLines = False
    title_merge(ws4, 'A1:D1', '📈  PROFIT & LOSS SUMMARY  |  BrokerNote AI')
    ws4.row_dimensions[1].height = 32
    ws4.column_dimensions['A'].width = 38
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 20

    # Section header helper
    def pnl_section(ws, row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=DARK_BLUE)
            ws.cell(row=row, column=col).border = make_border()
        c.border = make_border()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        return row + 1

    def pnl_row(ws, row, label, value, is_total=False, positive_green=True):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=4, value=round(value, 2) if value is not None else None)
        lc.border = vc.border = make_border()
        vc.number_format = '₹#,##0.00'
        if is_total:
            color = ("16A34A" if value >= 0 else "DC2626") if positive_green else DARK_BLUE
            lc.font = Font(bold=True, size=11, color=DARK_BLUE)
            vc.font = Font(bold=True, size=12, color=color)
            lc.fill = vc.fill = PatternFill("solid", fgColor="DBEAFE")
        else:
            lc.font = Font(size=10)
            vc.font = Font(size=10)
        return row + 1

    pr = 3  # start row

    # ── Equity Section
    total_buy_val  = sum(t.get('total_buy_value', 0)  for t in equity_trades)
    total_sell_val = sum(t.get('total_sell_value', 0) for t in equity_trades)
    equity_pnl     = total_sell_val - total_buy_val

    if equity_trades:
        pr = pnl_section(ws4, pr, '📈  EQUITY SEGMENT')
        pr = pnl_row(ws4, pr, 'Total Buy Value (all securities)', total_buy_val)
        pr = pnl_row(ws4, pr, 'Total Sell Value (all securities)', total_sell_val)
        pr = pnl_row(ws4, pr, 'Equity Gross P&L  (Sell − Buy)', equity_pnl, is_total=True)
        pr += 1

    # ── F&O Section
    fo_pnl = sum(d.get('net_total', 0) for d in derivatives)
    if derivatives:
        pr = pnl_section(ws4, pr, '📊  DERIVATIVE / F&O SEGMENT')
        for deriv in derivatives:
            label = f"  {deriv.get('buy_sell','')}  {deriv.get('contract','')[:35]}"
            pnl_val = deriv.get('net_total', 0)
            lc = ws4.cell(row=pr, column=1, value=label)
            vc = ws4.cell(row=pr, column=4, value=round(pnl_val, 2))
            lc.border = vc.border = make_border()
            lc.font = Font(size=10)
            vc.font = Font(size=10, color='16A34A' if pnl_val >= 0 else 'DC2626')
            vc.number_format = '₹#,##0.00'
            pr += 1
        pr = pnl_row(ws4, pr, 'Net F&O P&L', fo_pnl, is_total=True)
        pr += 1

    # ── Expenses Section
    total_expenses = stt + gst + exc + sebi + stamp + ipf
    pr = pnl_section(ws4, pr, '💰  TOTAL EXPENSES (All Segments Combined)')
    exp_detail = [
        ('Securities Transaction Tax (STT)', stt),
        ('GST Expense (CGST @ 9% + SGST @ 9%)', gst),
        ('Exchange Transaction Charges', exc),
        ('SEBI Turnover Fees', sebi),
        ('Stamp Duty', stamp),
        ('IPF Charges', ipf),
    ]
    for label, val in exp_detail:
        if val and val > 0:
            pr = pnl_row(ws4, pr, f'  {label}', val)
    pr = pnl_row(ws4, pr, 'Total Expenses', total_expenses, is_total=True, positive_green=False)
    pr += 1

    # ── Net P&L
    pr = pnl_section(ws4, pr, '🏁  NET SETTLEMENT')
    net_pnl_overall = equity_pnl + fo_pnl - total_expenses
    pr = pnl_row(ws4, pr, 'Equity Gross P&L', equity_pnl)
    if derivatives:
        pr = pnl_row(ws4, pr, 'F&O Net P&L', fo_pnl)
    pr = pnl_row(ws4, pr, 'Less: Total Expenses', -total_expenses)
    pr = pnl_row(ws4, pr, '★  NET P&L (after expenses)', net_pnl_overall, is_total=True)
    pr += 1
    pr = pnl_row(ws4, pr, 'Net Amount Payable / (Receivable) from Broker',
                 net, is_total=True, positive_green=False)

    # ITR filing note
    pr += 2
    note_c = ws4.cell(row=pr, column=1,
        value='ℹ  ITR Filing Note:  Equity gains → Schedule CG (STCG/LTCG)  |  '
              'F&O income → Schedule BP (Business Profession)  |  Expenses → Deductible')
    note_c.font = Font(italic=True, size=9, color='374151')
    note_c.fill = PatternFill("solid", fgColor="FEF9C3")
    ws4.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    contract_no = header_info.get('contract_note_no', 'ContractNote').replace('/', '-')
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=BrokerNote_{contract_no}.xlsx"}
    )


# ─── Tally XML — Balanced Journal Voucher ────────────────────────────────────

@app.post("/tally-xml")
async def generate_tally_xml(data: dict):
    header      = data.get('header', {})
    equity      = data.get('trades', [])
    derivatives = data.get('derivatives', [])
    charges     = data.get('charges', {})
    broker      = data.get('broker_info', {}).get('broker', 'Broker')

    trade_date = header.get('trade_date', '')
    try:
        d = datetime.datetime.strptime(trade_date, '%d/%m/%Y')
        tally_date = d.strftime('%Y%m%d')
    except Exception:
        tally_date = datetime.datetime.now().strftime('%Y%m%d')

    narration = (
        f"Contract Note: {header.get('contract_note_no', '')} | "
        f"Client: {header.get('client_name', '')} | PAN: {header.get('pan', '')} | "
        f"Broker: {broker} | Settlement: {header.get('settlement_no', '')}"
    )

    gst_val   = charges.get('gst',  round(charges.get('cgst', 0) + charges.get('sgst', 0), 2))
    stt_val   = charges.get('stt', 0)
    exc_val   = charges.get('exchange_charges', 0)
    sebi_val  = charges.get('sebi_fees', 0)
    stamp_val = charges.get('stamp_duty', 0)
    ipf_val   = charges.get('ipf_charges', 0)

    # Build balanced Dr / Cr entry lists
    # dr_list / cr_list = list of (ledger_name, amount)
    dr_list = []
    cr_list = []

    # Equity buys → Dr. Stock A/c
    for trade in equity:
        buy_val = trade.get('total_buy_value', 0)
        if buy_val > 0:
            dr_list.append((f"{trade.get('security', 'Investment')} A/c", buy_val))

    # Equity sells → Cr. Stock A/c
    for trade in equity:
        sell_val = trade.get('total_sell_value', 0)
        if sell_val > 0:
            cr_list.append((f"{trade.get('security', 'Investment')} A/c", sell_val))

    # F&O: loss → Dr., profit → Cr.
    for deriv in derivatives:
        net = deriv.get('net_total', 0)
        if net < 0:
            dr_list.append(('F&O Trading P&L A/c', abs(net)))
        elif net > 0:
            cr_list.append(('F&O Trading P&L A/c', net))

    # Expenses → Dr.
    for ledger, amt in [
        ('STT Expense A/c',          stt_val),
        ('GST Expense A/c',          gst_val),
        ('Exchange Charges Exp A/c', exc_val),
        ('SEBI Charges Exp A/c',     sebi_val),
        ('Stamp Duty Expense A/c',   stamp_val),
        ('IPF Charges Exp A/c',      ipf_val),
    ]:
        if amt and amt > 0:
            dr_list.append((ledger, amt))

    # Balancing entry: net broker payable / receivable
    total_dr = sum(a for _, a in dr_list)
    total_cr = sum(a for _, a in cr_list)
    diff = round(total_dr - total_cr, 2)
    if diff > 0:
        cr_list.append((f'{broker} Payable A/c', diff))
    elif diff < 0:
        dr_list.append((f'{broker} Receivable A/c', abs(diff)))

    # Build XML
    def xml_entry(ledger: str, is_dr: bool, amount: float) -> list:
        """ISDEEMEDPOSITIVE=Yes → Dr (amount negative); No → Cr (amount positive)."""
        sign = '-' if is_dr else ''
        return [
            '        <ALLLEDGERENTRIES.LIST>',
            f'          <LEDGERNAME>{ledger}</LEDGERNAME>',
            f'          <ISDEEMEDPOSITIVE>{"Yes" if is_dr else "No"}</ISDEEMEDPOSITIVE>',
            f'          <AMOUNT>{sign}{abs(amount):.2f}</AMOUNT>',
            '        </ALLLEDGERENTRIES.LIST>',
        ]

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<ENVELOPE>',
        '  <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>',
        '  <BODY><IMPORTDATA>',
        '    <REQUESTDESC>',
        '      <REPORTNAME>Vouchers</REPORTNAME>',
        '      <STATICVARIABLES>',
        '        <SVCURRENTCOMPANY>##SVCURRENTCOMPANY##</SVCURRENTCOMPANY>',
        '      </STATICVARIABLES>',
        '    </REQUESTDESC>',
        '    <REQUESTDATA>',
        '      <TALLYMESSAGE xmlns:UDF="TallyUDF">',
        '        <VOUCHER VCHTYPE="Journal" ACTION="Create" OBJVIEW="Accounting Voucher View">',
        f'          <DATE>{tally_date}</DATE>',
        f'          <NARRATION>{narration}</NARRATION>',
        '          <VOUCHERTYPENAME>Journal</VOUCHERTYPENAME>',
    ]

    for ledger, amt in dr_list:
        lines += xml_entry(ledger, is_dr=True,  amount=amt)
    for ledger, amt in cr_list:
        lines += xml_entry(ledger, is_dr=False, amount=amt)

    lines += [
        '        </VOUCHER>',
        '      </TALLYMESSAGE>',
        '    </REQUESTDATA>',
        '  </IMPORTDATA></BODY>',
        '</ENVELOPE>',
    ]

    xml_bytes = '\n'.join(lines).encode('utf-8')
    contract_no = header.get('contract_note_no', 'ContractNote').replace('/', '-')
    return StreamingResponse(
        io.BytesIO(xml_bytes),
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename=Tally_{contract_no}.xml"}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# CA COMPLIANCE CALENDAR — Routes
# ═══════════════════════════════════════════════════════════════════════════════

import os
from fastapi import Header as FHeader

CA_SENDER = os.getenv("CA_SENDER_NAME", "DigiAgentix CA Tools")
TWILIO_SID   = os.getenv("TWILIO_ACCOUNT_SID", "")
TWILIO_TOKEN = os.getenv("TWILIO_AUTH_TOKEN", "")
TWILIO_FROM  = os.getenv("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")


def ca_get_db():
    return sqlite3.connect('/tmp/brokernote.db')


def ca_auth(authorization: str) -> int:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization required")
    try:
        token = authorization.split(" ", 1)[1]
        email = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])["email"]
        conn  = ca_get_db()
        row   = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        conn.close()
        if not row: raise Exception()
        return row[0]
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def ca_send_whatsapp(mobile: str, message: str) -> tuple:
    """Returns (success: bool, error: str)"""
    sid   = os.getenv("TWILIO_ACCOUNT_SID", "")
    token = os.getenv("TWILIO_AUTH_TOKEN", "")
    frm   = os.getenv("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
    if not sid or not token:
        return False, "Twilio credentials not set in environment variables"
    try:
        from twilio.rest import Client as TwilioClient
        to = f"whatsapp:{mobile}" if not mobile.startswith("whatsapp:") else mobile
        TwilioClient(sid, token).messages.create(from_=frm, to=to, body=message)
        return True, ""
    except Exception as e:
        print(f"Twilio error: {e}")
        return False, str(e)


def ca_build_msg(name: str, filing: str, due: str) -> str:
    try:
        d = datetime.datetime.strptime(due, "%Y-%m-%d").strftime("%d %b %Y")
    except Exception:
        d = due
    return (f"Dear {name},\n\nReminder: *{filing}* filing is due on *{d}*.\n"
            f"Please share the required documents at the earliest.\n\n— {CA_SENDER}")


def ca_row_to_dict(row) -> dict:
    return {"id": row[0], "name": row[1], "pan": row[2] or "",
            "mobile": row[3], "filing_type": row[4], "due_date": row[5], "status": row[6]}


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/ca/auth/signup")
async def ca_signup(body: dict):
    email = body.get("email", "").lower().strip()
    pw    = body.get("password", "")
    name  = body.get("name", "").strip()
    firm  = body.get("firm", "").strip()
    if not email or not pw or not name:
        raise HTTPException(400, "Email, password and name required")
    if len(pw) < 6:
        raise HTTPException(400, "Password min 6 characters")
    conn = ca_get_db()
    try:
        conn.execute(
            "INSERT INTO users (email,password,name,firm,created_at) VALUES (?,?,?,?,?)",
            (email, hash_password(pw), name, firm, datetime.datetime.now().isoformat())
        )
        conn.commit()
        return {"token": create_token(email), "email": email, "name": name, "firm": firm}
    except sqlite3.IntegrityError:
        raise HTTPException(400, "Email already registered")
    finally:
        conn.close()


@app.post("/ca/auth/login")
async def ca_login(body: dict):
    email = body.get("email", "").lower().strip()
    pw    = body.get("password", "")
    conn  = ca_get_db()
    row   = conn.execute(
        "SELECT email,name,firm FROM users WHERE email=? AND password=?",
        (email, hash_password(pw))
    ).fetchone()
    conn.close()
    if not row:
        raise HTTPException(401, "Invalid email or password")
    return {"token": create_token(row[0]), "email": row[0], "name": row[1], "firm": row[2]}


# ── Clients CRUD ──────────────────────────────────────────────────────────────

@app.get("/ca/clients")
async def ca_get_clients(authorization: str = FHeader(None)):
    uid  = ca_auth(authorization)
    conn = ca_get_db()
    rows = conn.execute(
        "SELECT id,name,pan,mobile,filing_type,due_date,status FROM ca_clients WHERE user_id=? ORDER BY due_date ASC",
        (uid,)
    ).fetchall()
    conn.close()
    return [ca_row_to_dict(r) for r in rows]


@app.post("/ca/clients")
async def ca_add_client(body: dict, authorization: str = FHeader(None)):
    uid  = ca_auth(authorization)
    name, mobile = body.get("name","").strip(), body.get("mobile","").strip()
    filing, due  = body.get("filing_type","").strip(), body.get("due_date","").strip()
    pan, status  = body.get("pan","").strip(), body.get("status","Pending")
    if not name or not mobile or not filing or not due:
        raise HTTPException(400, "name, mobile, filing_type, due_date required")
    conn = ca_get_db()
    cur  = conn.execute(
        "INSERT INTO ca_clients (user_id,name,pan,mobile,filing_type,due_date,status,created_at) VALUES (?,?,?,?,?,?,?,?)",
        (uid, name, pan, mobile, filing, due, status, datetime.datetime.now().isoformat())
    )
    conn.commit()
    row = conn.execute("SELECT id,name,pan,mobile,filing_type,due_date,status FROM ca_clients WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return ca_row_to_dict(row)


@app.put("/ca/clients/{client_id}")
async def ca_update_client(client_id: int, body: dict, authorization: str = FHeader(None)):
    uid  = ca_auth(authorization)
    conn = ca_get_db()
    if not conn.execute("SELECT id FROM ca_clients WHERE id=? AND user_id=?", (client_id, uid)).fetchone():
        conn.close(); raise HTTPException(404, "Not found")
    fields = {k: v for k, v in body.items() if k in ("name","pan","mobile","filing_type","due_date","status")}
    if fields:
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE ca_clients SET {sets} WHERE id=?", (*fields.values(), client_id))
        conn.commit()
    row = conn.execute("SELECT id,name,pan,mobile,filing_type,due_date,status FROM ca_clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    return ca_row_to_dict(row)


@app.delete("/ca/clients/{client_id}")
async def ca_delete_client(client_id: int, authorization: str = FHeader(None)):
    uid = ca_auth(authorization)
    conn = ca_get_db()
    conn.execute("DELETE FROM ca_clients WHERE id=? AND user_id=?", (client_id, uid))
    conn.commit(); conn.close()
    return {"deleted": True}


# ── Reminders ─────────────────────────────────────────────────────────────────

@app.post("/ca/remind/test/{client_id}")
async def ca_test_reminder(client_id: int, authorization: str = FHeader(None)):
    uid  = ca_auth(authorization)
    conn = ca_get_db()
    row  = conn.execute(
        "SELECT name,mobile,filing_type,due_date FROM ca_clients WHERE id=? AND user_id=?",
        (client_id, uid)
    ).fetchone()
    conn.close()
    if not row: raise HTTPException(404, "Client not found")
    name, mobile, filing, due = row
    msg = ca_build_msg(name, filing, due)
    ok, err = ca_send_whatsapp(mobile, msg)
    result = "sent-manual" if ok else "failed"
    conn2 = ca_get_db()
    conn2.execute(
        "INSERT INTO ca_message_log (user_id,client_name,filing_type,due_date,mobile,sent_at,result) VALUES (?,?,?,?,?,?,?)",
        (uid, name, filing, due, mobile, datetime.datetime.now().isoformat(), result)
    )
    conn2.commit(); conn2.close()
    return {"success": ok, "message": msg, "to": mobile, "error": err}


@app.get("/ca/debug")
async def ca_debug():
    sid   = os.getenv("TWILIO_ACCOUNT_SID", "")
    token = os.getenv("TWILIO_AUTH_TOKEN", "")
    frm   = os.getenv("TWILIO_WHATSAPP_FROM", "")
    return {
        "TWILIO_ACCOUNT_SID":    f"{sid[:6]}..." if sid else "NOT SET",
        "TWILIO_AUTH_TOKEN":     "SET" if token else "NOT SET",
        "TWILIO_WHATSAPP_FROM":  frm if frm else "NOT SET",
    }


@app.get("/ca/logs")
async def ca_get_logs(authorization: str = FHeader(None)):
    uid  = ca_auth(authorization)
    conn = ca_get_db()
    rows = conn.execute(
        "SELECT client_name,filing_type,due_date,mobile,sent_at,result FROM ca_message_log WHERE user_id=? ORDER BY sent_at DESC LIMIT 100",
        (uid,)
    ).fetchall()
    conn.close()
    return [{"client_name": r[0], "filing_type": r[1], "due_date": r[2],
             "mobile": r[3], "sent_at": r[4], "result": r[5]} for r in rows]
